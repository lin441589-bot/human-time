"""
独立程序：两行表头 + 成本公式（改进版）
- 固定列保持原样（垂直合并）
- 项目部分做成两行表头：项目名合并4列 + 子表头
- 尽量不破坏原始数据
- 只保留「总表」
"""

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils.cell import column_index_from_string
from copy import copy
import re

# ==================== 配置区 ====================
input_file  = '月度工时大表_含总表.xlsx'
output_file = '月度工时大表_含公式_两行表头.xlsx'
# ================================================

thin_border = Border(left=Side('thin'), right=Side('thin'), top=Side('thin'), bottom=Side('thin'))
center = Alignment(horizontal='center', vertical='center', wrap_text=True)
header_fill = PatternFill('solid', fgColor='D9E1F2')
total_fill = PatternFill('solid', fgColor='FFFF99')
bold = Font(bold=True)
normal = Font(bold=False)


def process_sheet(ws):
    print(f"  开始处理：{ws.title}")

    # ---------- 0. 取消所有合并 ----------
    merged_ranges = list(ws.merged_cells.ranges)
    for mr in merged_ranges:
        ws.unmerge_cells(str(mr))
    print(f"    已取消 {len(merged_ranges)} 个合并单元格")

    # ---------- 1. 找表头行 ----------
    header_row = None
    for r in range(1, 30):
        val = ws.cell(row=r, column=1).value
        if val and str(val).strip() == '员工':
            header_row = r
            break
    if header_row is None:
        print("    未找到「员工」表头，跳过")
        return

    # ---------- 2. 读取原表头 ----------
    old_headers = []
    c = 1
    while True:
        v = ws.cell(row=header_row, column=c).value
        if v is None or str(v).strip() == '':
            break
        old_headers.append(str(v).strip())
        c += 1
    print(f"    原表头列数：{len(old_headers)}")

    # ---------- 3. 找合计行 ----------
    total_row = None
    for r in range(header_row + 1, ws.max_row + 2):
        if ws.cell(row=r, column=1).value == '合计':
            total_row = r
            break
    if total_row is None:
        print("    未找到「合计」行，跳过")
        return

    first_data = header_row + 1
    last_data = total_row - 1
    n_data = last_data - first_data + 1
    print(f"    数据行 {first_data}~{last_data}（{n_data}人），合计行 {total_row}")

    # ---------- 4. 固定列 ----------
    fixed_names = ['员工', '基本工资', '加班工资', '总工资', '工时工资',
                   '出勤天数', '基本工时', '加班工时', '总工时']
    fixed_names = [n for n in fixed_names if n in old_headers]

    # ---------- 5. 读取固定列全部数据（含合计） ----------
    fixed_data = {}
    for name in fixed_names:
        col_idx = old_headers.index(name) + 1
        vals = [ws.cell(row=r, column=col_idx).value for r in range(first_data, total_row + 1)]
        fixed_data[name] = vals

    # ---------- 6. 读取所有项目工时 ----------
    projects = {}
    project_order = []
    for h in old_headers:
        m = re.match(r'^(.+?)(基本工时|加班工时)$', h)
        if not m:
            continue
        p, kind = m.group(1), m.group(2)
        if p not in projects:
            projects[p] = {'基本工时': [0.0]*n_data, '加班工时': [0.0]*n_data}
            project_order.append(p)
        col_idx = old_headers.index(h) + 1
        for i, r in enumerate(range(first_data, last_data + 1)):
            val = ws.cell(row=r, column=col_idx).value
            try:
                projects[p][kind][i] = float(val) if val is not None else 0.0
            except:
                projects[p][kind][i] = 0.0

    print(f"    识别到 {len(project_order)} 个项目")

    # ---------- 7. 备份标题区 ----------
    title_backup = []
    for r in range(1, header_row):
        row_data = []
        for c in range(1, 12):
            cell = ws.cell(row=r, column=c)
            row_data.append({
                'value': cell.value,
                'font': copy(cell.font),
                'alignment': copy(cell.alignment),
                'fill': copy(cell.fill),
            })
        title_backup.append(row_data)

    # ---------- 8. 清空数据区 ----------
    max_col = max(len(old_headers) + 5, 80)
    for r in range(header_row, total_row + 5):
        for c in range(1, max_col + 1):
            cell = ws.cell(row=r, column=c)
            cell.value = None
            cell.border = Border()
            cell.fill = PatternFill()
            cell.font = Font()
            cell.alignment = Alignment()

    # ---------- 9. 写回标题区 ----------
    for r_idx, row_data in enumerate(title_backup):
        for c_idx, info in enumerate(row_data):
            cell = ws.cell(row=r_idx+1, column=c_idx+1)
            cell.value = info['value']
            if info['font']: cell.font = info['font']
            if info['alignment']: cell.alignment = info['alignment']
            if info['fill']: cell.fill = info['fill']

    # ---------- 10. 写两行表头 ----------
    # 固定列（上下合并）
    for idx, name in enumerate(fixed_names, 1):
        cell = ws.cell(row=header_row, column=idx, value=name)
        cell.font = bold
        cell.alignment = center
        cell.border = thin_border
        cell.fill = header_fill

        cell2 = ws.cell(row=header_row+1, column=idx)
        cell2.border = thin_border
        cell2.fill = header_fill
        ws.merge_cells(start_row=header_row, start_column=idx,
                       end_row=header_row+1, end_column=idx)

    # 项目列（每个项目4列）
    col = len(fixed_names) + 1
    for p in project_order:
        # 项目名合并4列
        cell = ws.cell(row=header_row, column=col, value=p)
        cell.font = bold
        cell.alignment = center
        cell.border = thin_border
        cell.fill = header_fill
        for i in range(1, 4):
            c = ws.cell(row=header_row, column=col+i)
            c.border = thin_border
            c.fill = header_fill
        ws.merge_cells(start_row=header_row, start_column=col,
                       end_row=header_row, end_column=col+3)

        # 子表头
        for i, sub in enumerate(['基本工时', '加班工时', '基本成本', '加班成本']):
            cell = ws.cell(row=header_row+1, column=col+i, value=sub)
            cell.font = normal
            cell.alignment = center
            cell.border = thin_border
            cell.fill = header_fill
        col += 4

    # ---------- 11. 写回固定列数据 ----------
    new_first = header_row + 2
    new_total = new_first + n_data

    for idx, name in enumerate(fixed_names, 1):
        for i, v in enumerate(fixed_data[name]):
            r = new_first + i
            cell = ws.cell(row=r, column=idx, value=v)
            cell.alignment = center
            cell.border = thin_border
            if i == len(fixed_data[name]) - 1:
                cell.font = bold
                cell.fill = total_fill

    # ---------- 12. 写项目数据 + 成本公式 ----------
    wage_col = get_column_letter(fixed_names.index('工时工资') + 1)
    col = len(fixed_names) + 1

    for p in project_order:
        basic_vals = projects[p]['基本工时']
        ot_vals = projects[p]['加班工时']
        basic_l = get_column_letter(col)
        ot_l = get_column_letter(col + 1)

        for i in range(n_data):
            r = new_first + i
            # 基本工时
            cell = ws.cell(row=r, column=col, value=basic_vals[i])
            cell.alignment = center
            cell.border = thin_border
            # 加班工时
            cell = ws.cell(row=r, column=col+1, value=ot_vals[i])
            cell.alignment = center
            cell.border = thin_border
            # 基本成本
            cell = ws.cell(row=r, column=col+2,
                           value=f'=IF({wage_col}{r}="","",{basic_l}{r}*{wage_col}{r})')
            cell.alignment = center
            cell.border = thin_border
            # 加班成本
            cell = ws.cell(row=r, column=col+3,
                           value=f'=IF({wage_col}{r}="","",{ot_l}{r}*{wage_col}{r})')
            cell.alignment = center
            cell.border = thin_border

        # 合计行
        for i in range(4):
            letter = get_column_letter(col + i)
            cell = ws.cell(row=new_total, column=col+i,
                           value=f'=SUM({letter}{new_first}:{letter}{new_total-1})')
            cell.font = bold
            cell.fill = total_fill
            cell.alignment = center
            cell.border = thin_border

        for i in range(4):
            ws.column_dimensions[get_column_letter(col+i)].width = 11
        col += 4

    # ---------- 13. 总工资 / 工时工资公式 ----------
    b_sal = get_column_letter(fixed_names.index('基本工资') + 1)
    o_sal = get_column_letter(fixed_names.index('加班工资') + 1)
    t_sal = get_column_letter(fixed_names.index('总工资') + 1)
    wage = get_column_letter(fixed_names.index('工时工资') + 1)
    t_hour = get_column_letter(fixed_names.index('总工时') + 1)

    for i in range(n_data):
        r = new_first + i
        ws[f'{t_sal}{r}'] = f'=IF(OR({b_sal}{r}="",{o_sal}{r}=""),"",{b_sal}{r}+{o_sal}{r})'
        ws[f'{wage}{r}'] = f'=IF(OR({t_sal}{r}="",{t_hour}{r}=0),"",{t_sal}{r}/{t_hour}{r})'

    print(f"    完成！新数据行 {new_first}~{new_total-1}，合计行 {new_total}")
    print(f"    项目数：{len(project_order)}")


# ==================== 主程序 ====================
print(f"打开：{input_file}")
wb = load_workbook(input_file)

# 只保留总表
for name in list(wb.sheetnames):
    if name != '总表':
        print(f"删除：{name}")
        del wb[name]

for name in wb.sheetnames:
    process_sheet(wb[name])

wb.save(output_file)
print(f"\n全部完成！输出：{output_file}")