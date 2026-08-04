import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter
from datetime import datetime

# ==================== 配置区 ====================
input_file = 'man-hours.xlsx'
output_file = '月度工时大表_含总表.xlsx'
# ================================================


def create_monthly_df(df_input):
    """生成 monthly 数据（含合计行）—— 总工时也排除 G300"""
    
    # ==================== 出勤 + 总工时 都排除 G300 ====================
    df_non_g300 = df_input[df_input['项目号'] != 'G300'].copy()
    
    # 出勤天数（非 G300 有工时的日期）
    daily_total = df_non_g300.groupby(['员工', '日期'])['总工时'].sum().reset_index()
    attendance = daily_total[daily_total['总工时'] > 0].groupby('员工')['日期'].nunique().reset_index()
    attendance.rename(columns={'日期': '出勤天数'}, inplace=True)

    # 基本工时 / 加班工时 / 总工时 也排除 G300
    emp_summary = df_non_g300.groupby('员工').agg({
        '基本工时': 'sum',
        '加班工时': 'sum',
        '总工时': 'sum'
    }).reset_index()

    emp_summary = pd.merge(emp_summary, attendance, on='员工', how='left')
    emp_summary['出勤天数'] = emp_summary['出勤天数'].fillna(0).astype(int)

    emp_summary['基本工资'] = None
    emp_summary['加班工资'] = None
    emp_summary['总工资'] = None
    emp_summary['工时工资'] = None

    # 透视项目工时
    basic_pivot = df_input.pivot_table(index='员工', columns='项目号', values='基本工时', aggfunc='sum', fill_value=0)
    overtime_pivot = df_input.pivot_table(index='员工', columns='项目号', values='加班工时', aggfunc='sum', fill_value=0)

    projects = sorted(set(basic_pivot.columns.union(overtime_pivot.columns)))
    for p in ['G100', 'G200']:
        if p not in projects:
            projects.append(p)
    if 'G300' in projects:
        projects.remove('G300')
    projects.sort()

    project_columns = []
    for p in projects:
        project_columns.append(f'{p}基本工时')
        project_columns.append(f'{p}加班工时')

    proj_df = pd.DataFrame(0.0, index=emp_summary.set_index('员工').index, columns=project_columns)

    for p in projects:
        if p in basic_pivot.columns:
            proj_df[f'{p}基本工时'] = basic_pivot[p]
        if p in overtime_pivot.columns:
            proj_df[f'{p}加班工时'] = overtime_pivot[p]

    monthly = pd.concat([emp_summary.set_index('员工'), proj_df], axis=1).reset_index()

    column_order = ['员工', '基本工资', '加班工资', '总工资', '工时工资',
                    '出勤天数', '基本工时', '加班工时', '总工时'] + project_columns

    monthly = monthly[column_order]
    monthly = monthly.sort_values('员工').reset_index(drop=True)

    # 合计行
    numeric_cols = monthly.select_dtypes(include='number').columns
    total_row = monthly[numeric_cols].sum(numeric_only=True)
    total_row['员工'] = '合计'
    total_row['基本工资'] = None
    total_row['加班工资'] = None
    total_row['总工资'] = None
    total_row['工时工资'] = None
    total_row['出勤天数'] = ''

    monthly = pd.concat([monthly, pd.DataFrame([total_row])], ignore_index=True)

    return monthly, project_columns


def write_sheet(ws, monthly_df, project_columns, year, month_name_cn, start_date, end_date, dept_name=None):
    """向工作表写入格式化内容"""
    bold_font = Font(bold=True)
    normal_font = Font(bold=False)
    center_align = Alignment(horizontal='center', vertical='center')
    total_fill = PatternFill('solid', fgColor='FFFF99')
    thin_border = Border(
        left=Side('thin'), right=Side('thin'),
        top=Side('thin'), bottom=Side('thin')
    )

    # 标题
    title = f"Man Hours for Each Project ({dept_name + ' - ' if dept_name else ''}{start_date} to {end_date})"
    ws['A1'] = title
    ws.merge_cells('A1:D1')
    ws['A1'].font = bold_font
    ws['A1'].alignment = center_align

    ws['A2'] = "公司: "
    ws['A3'] = "年份"
    ws['B3'] = year
    ws['C3'] = "月份"
    ws['D3'] = month_name_cn

    ws['A4'] = "當月上班天數（大小周）"
    ws['B4'] = ""
    ws['A5'] = "當月上班總基本工時"
    ws['B5'] = ""

    ws['A6'] = '*按"基本工资/（基本工時+加班工時）"计算平均工时工資'
    ws['A7'] = "注：G300为休息日，不计入出勤统计，亦不显示工时列"

    for row_num in range(2, 8):
        ws[f'A{row_num}'].font = bold_font
        ws[f'A{row_num}'].alignment = center_align
        if row_num in [3, 4, 5]:
            ws[f'B{row_num}'].font = bold_font
            ws[f'B{row_num}'].alignment = center_align
        if row_num == 3:
            for col in ['C', 'D']:
                ws[f'{col}{row_num}'].font = bold_font
                ws[f'{col}{row_num}'].alignment = center_align

    # 数据区
    data_start_row = 8
    for r in dataframe_to_rows(monthly_df, index=False, header=True):
        ws.append(r)

    # 表头样式
    for cell in ws[data_start_row]:
        cell.font = normal_font
        cell.alignment = center_align
        cell.border = thin_border

    # 数据行 & 合计行样式
    for row in ws.iter_rows(min_row=data_start_row + 1, max_row=ws.max_row):
        is_total_row = ws.cell(row[0].row, 1).value == '合计'
        for cell in row:
            cell.border = thin_border
            cell.alignment = center_align
            if is_total_row:
                cell.font = bold_font
                cell.fill = total_fill
            else:
                cell.font = normal_font

    # 工资统计区
    summary_start_row = ws.max_row + 3
    ws[f'A{summary_start_row}'] = "辦公室員工工資："
    ws[f'A{summary_start_row + 1}'] = "其它差額："
    ws[f'A{summary_start_row + 2}'] = "總工資："

    salary_col = 'E'
    ws[f'{salary_col}{summary_start_row}'] = "基本工資合計："
    ws[f'{salary_col}{summary_start_row + 1}'] = "加班工資合計："
    ws[f'{salary_col}{summary_start_row + 3}'] = "工資及員工費用合計(1+14%):"

    for r in range(summary_start_row, summary_start_row + 4):
        if ws[f'A{r}'].value:
            ws[f'A{r}'].font = bold_font
            ws[f'A{r}'].alignment = center_align
        if ws[f'{salary_col}{r}'].value:
            ws[f'{salary_col}{r}'].font = bold_font
            ws[f'{salary_col}{r}'].alignment = center_align

    # 备注
    note_row = summary_start_row + 5
    ws[f'A{note_row}'] = "备注："
    ws[f'A{note_row}'].font = bold_font
    ws.merge_cells(f'A{note_row}:C{note_row}')
    ws[f'A{note_row}'].alignment = center_align

    ws[f'A{note_row + 1}'] = "1. 总工时由基本工时和加班工时构成（已排除G300）。"
    ws[f'A{note_row + 2}'] = "2. 表上“基本工资”，“加班工资”欄需由计工资的同事录入"
    ws[f'A{note_row + 3}'] = "3. 上班天数及总基本工时请自行填写"

    # 自动调整列宽
    for col_cells in ws.columns:
        max_length = 0
        column_letter = None
        for cell in col_cells:
            if cell.value is not None:
                max_length = max(max_length, len(str(cell.value)))
            if column_letter is None and hasattr(cell, 'column'):
                column_letter = get_column_letter(cell.column)
        if column_letter:
            ws.column_dimensions[column_letter].width = min(max_length + 2, 40)


# ==================== 主程序 ====================
df = pd.read_excel(input_file, sheet_name='values (2)', header=0)
df.columns = [col.strip() for col in df.columns]

for col in df.columns:
    if df[col].apply(lambda x: isinstance(x, list)).any():
        df[col] = df[col].apply(lambda x: x[0] if isinstance(x, list) and len(x) > 0 else x)
    if col not in ['日期', '基本工时', '加班工时']:
        df[col] = df[col].astype(str).str.strip()

df['员工'] = df['员工'].str.strip()
df['项目号'] = df['项目号'].str.strip()
df['基本工时'] = pd.to_numeric(df['基本工时'], errors='coerce').fillna(0)
df['加班工时'] = pd.to_numeric(df['加班工时'], errors='coerce').fillna(0)
df['总工时'] = df['基本工时'] + df['加班工时']

# ========== 新增：统一部门名称 ==========
df['部门'] = df['部门'].replace('工场部', '工厂部')
# ========================================

if df['日期'].dtype in ['int64', 'float64']:
    df['日期'] = pd.to_datetime(df['日期'], unit='D', origin='1899-12-30')
else:
    df['日期'] = pd.to_datetime(df['日期'])

if not df.empty:
    year = int(df['日期'].dt.year.mode()[0])
    month = int(df['日期'].dt.month.mode()[0])
else:
    year = datetime.now().year
    month = datetime.now().month

month_name_cn = f'{month}月'
start_date = df['日期'].min().strftime('%Y/%m/%d') if not df.empty else ""
end_date = df['日期'].max().strftime('%Y/%m/%d') if not df.empty else ""

wb = Workbook()
wb.remove(wb.active)

# 总表
monthly_total, project_cols_total = create_monthly_df(df)
ws_total = wb.create_sheet(title='总表')
write_sheet(ws_total, monthly_total, project_cols_total, year, month_name_cn, start_date, end_date)

# 部门表
departments = {
    '技术部': '技术部',
    '工厂部': '工厂部',
}

for dept_key, dept_name in departments.items():
    df_dept = df[df['部门'].str.contains(dept_name, na=False)].copy()
    if df_dept.empty:
        print(f"警告：未找到 {dept_name} 的任何数据，跳过该部门")
        continue

    monthly_dept, _ = create_monthly_df(df_dept)
    ws_dept = wb.create_sheet(title=dept_name)
    write_sheet(ws_dept, monthly_dept, project_cols_total, year, month_name_cn, start_date, end_date, dept_name=dept_name)

wb.save(output_file)

print(f"已生成文件：{output_file}")
print(f"包含工作表：总表 + {', '.join([name for name in departments.values() if df[df['部门'].str.contains(name, na=False)].shape[0] > 0])}")
print("上班天数及总基本工时已留空，请打开 Excel 自行填写")
print("公司名称请在 B2 单元格自行填写")
print("已自动把「工场部」替换为「工厂部」")
print("已修复：总工时现在也排除 G300，与项目工时之和一致")